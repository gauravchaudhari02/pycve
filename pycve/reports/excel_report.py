"""Excel report generator using openpyxl."""

from __future__ import annotations

from collections import Counter
from datetime import datetime
from pathlib import Path

from pycve.models.cve import CVERecord
from pycve.utils.exceptions import MissingDependencyError

# Severity → hex fill colours
_SEV_FILLS = {
    "CRITICAL": "DC3545",
    "HIGH":     "FD7E14",
    "MEDIUM":   "FFC107",
    "LOW":      "198754",
    "UNKNOWN":  "6C757D",
}


def generate_excel_report(cves: list[CVERecord], output_path: str | Path) -> str:
    """Generate a multi-sheet Excel report.

    Sheets:  Summary, CVE Details, References, Patches

    Returns the absolute path of the created file.
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise MissingDependencyError("openpyxl", "Excel reports", "reports")

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()

    # ── Summary sheet ─────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum["A1"] = "PyCVE Security Report"
    ws_sum["A1"].font = Font(bold=True, size=16, color="6366F1")
    ws_sum["A2"] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
    ws_sum["A3"] = f"Total CVEs: {len(cves)}"
    ws_sum["A5"] = "Severity Distribution"
    ws_sum["A5"].font = Font(bold=True)

    dist = Counter(cve.severity for cve in cves)
    row = 6
    for sev in ["CRITICAL", "HIGH", "MEDIUM", "LOW", "UNKNOWN"]:
        count = dist.get(sev, 0)
        ws_sum.cell(row, 1, sev).font = Font(color=_SEV_FILLS.get(sev, "000000"), bold=True)
        ws_sum.cell(row, 2, count)
        row += 1

    ws_sum["A12"] = "Patched CVEs"
    ws_sum["B12"] = sum(1 for cve in cves if cve.patch_references)
    ws_sum["A13"] = "Unpatched CVEs"
    ws_sum["B13"] = len(cves) - sum(1 for cve in cves if cve.patch_references)
    ws_sum.column_dimensions["A"].width = 22
    ws_sum.column_dimensions["B"].width = 12

    # ── CVE Details sheet ─────────────────────────────────────────────────
    ws = wb.create_sheet("CVE Details")
    headers = ["CVE ID", "Severity", "CVSS Score", "Published", "Modified", "Status", "Description", "CWEs"]
    _write_header_row(ws, headers)

    for cve in cves:
        cwes = ", ".join(w.cwe_id for w in cve.weaknesses) or "—"
        fill = PatternFill("solid", fgColor=_SEV_FILLS.get(cve.severity, "6C757D") + "33")
        row_values = [
            cve.id,
            cve.severity,
            cve.cvss_score,
            cve.published.strftime("%Y-%m-%d") if cve.published else None,
            cve.last_modified.strftime("%Y-%m-%d") if cve.last_modified else None,
            cve.vuln_status,
            cve.description,
            cwes,
        ]
        r = ws.max_row + 1
        for col, val in enumerate(row_values, 1):
            cell = ws.cell(r, col, val)
            if col == 2:  # severity cell
                cell.fill = fill
                cell.font = Font(bold=True)
            if col == 7:  # description cell
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = None  # auto-height

    _auto_width(ws, desc_col=7)
    ws.auto_filter.ref = ws.dimensions

    # ── References sheet ──────────────────────────────────────────────────
    ws_ref = wb.create_sheet("References")
    _write_header_row(ws_ref, ["CVE ID", "URL", "Tags"])
    for cve in cves:
        for ref in cve.references:
            ws_ref.append([cve.id, ref.url, ", ".join(ref.tags)])
    _auto_width(ws_ref)

    # ── Patches sheet ─────────────────────────────────────────────────────
    ws_pat = wb.create_sheet("Patches")
    _write_header_row(ws_pat, ["CVE ID", "Patch URL"])
    for cve in cves:
        for ref in cve.patch_references:
            ws_pat.append([cve.id, ref.url])
    _auto_width(ws_pat)

    wb.save(str(output_path))
    return str(output_path.resolve())


def _write_header_row(ws, headers: list[str]) -> None:
    from openpyxl.styles import Font, PatternFill
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1A1D26")


def _auto_width(ws, desc_col: int | None = None) -> None:
    from openpyxl.utils import get_column_letter
    for col in ws.columns:
        col_idx = col[0].column
        if desc_col is not None and col_idx == desc_col:
            ws.column_dimensions[get_column_letter(col_idx)].width = 60
        else:
            max_len = max((len(str(c.value or "")) for c in col), default=0)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)
